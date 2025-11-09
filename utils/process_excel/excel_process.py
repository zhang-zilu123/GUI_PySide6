import xlrd
import openpyxl
import excel2img

from openpyxl import Workbook

# 将.xls文件转换为.xlsx格式
def convert_xls_to_xlsx(input_file, output_file):
    """
    将.xls文件转换为.xlsx格式。

    参数：
        input_file (str): .xls文件的路径。
        output_file (str): 保存转换后.xlsx文件的路径。

    返回：
        None
    """
    # 打开.xls文件
    workbook_xls = xlrd.open_workbook(input_file)
    workbook_xlsx = Workbook()

    for sheet_index in range(workbook_xls.nsheets):
        sheet_xls = workbook_xls.sheet_by_index(sheet_index)
        sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)

        for row_idx in range(sheet_xls.nrows):
            row_values = sheet_xls.row_values(row_idx)
            sheet_xlsx.append(row_values)

    # 删除openpyxl创建的默认工作表
    if "Sheet" in workbook_xlsx.sheetnames:
        del workbook_xlsx["Sheet"]

    # 保存新的.xlsx文件
    workbook_xlsx.save(output_file)
    print(f"已将 '{input_file}' 转换为 '{output_file}'")

# 将Excel文件拆分为多个工作表
def split_excel_sheets(input_file, output_dir):
    """
    将一个Excel文件拆分为多个工作表，并将每个工作表保存为一个新的Excel文件。

    参数：
        input_file (str): 输入Excel文件的路径。
        output_dir (str): 保存拆分工作表的目录。

    返回：
        None
    """
    import os

    # 加载工作簿
    try:
        workbook = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"加载工作簿时出错: {e}")
        return

    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 遍历工作簿中的每个工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 为当前工作表创建一个新的工作簿
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = sheet_name

        # 将原始工作表的数据复制到新工作表
        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # 保存新的工作簿
        output_file = os.path.join(output_dir, f"{sheet_name}.xlsx")
        try:
            new_workbook.save(output_file)
            print(f"已将工作表 '{sheet_name}' 保存到 '{output_file}'")
        except Exception as e:
            print(f"保存工作表 '{sheet_name}' 时出错: {e}")

# 将Excel文件转换为图片
def convert_excel_to_images(file_paths, output_dir):
    """
    使用excel2img包将给定的Excel文件列表转换为图片。

    参数：
        file_paths (list): 包含Excel文件路径的列表。
        output_dir (str): 保存图片的目录。

    返回：
        None
    """
    import os

    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 遍历文件路径列表中的所有Excel文件
    for input_file in file_paths:
        if input_file.endswith(".xlsx"):
            output_file = os.path.join(
                output_dir, f"{os.path.splitext(os.path.basename(input_file))[0]}.png"
            )

            try:
                # 将Excel转换为图片
                excel2img.export_img(input_file, output_file)
                print(f"已将 '{input_file}' 转换为 '{output_file}'")
            except Exception as e:
                print(f"转换 '{input_file}' 时出错: {e}")

if __name__ == "__main__":
    # 示例用法
    input_file = "../布局1_扁平式布局.xls"  # 替换为你的输入文件路径
    # output_dir = "output_sheets"  # 替换为你想要的输出目录
    # convert_xls_to_xlsx(input_file, "output.xlsx")
    # input_file = "./output.xlsx"
    # split_excel_sheets(input_file, output_dir)
    input_files = [
        "./布局2_主表+子表布局.xls",
        "./布局3_分块布局.xls",
    ]
    convert_xls_to_xlsx(input_files[0], "../布局2_主表+子表布局.xlsx")
    convert_xls_to_xlsx(input_files[1], "../布局3_分块布局.xlsx")
    input_files = [
        "./布局2_主表+子表布局.xlsx",
        "./布局3_分块布局.xlsx",
    ]
    output_dir = "../output_images"  # 替换为你想要的输出目录
    convert_excel_to_images(input_files, output_dir)
