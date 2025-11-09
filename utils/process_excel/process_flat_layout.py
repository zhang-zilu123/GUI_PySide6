import json

import openpyxl
import os
import dashscope
from dashscope import Generation
from dotenv import load_dotenv
from openpyxl.styles import Border, Side

from config.config import HEADER_ROW_DETECTION_PROMPT
from utils.process_excel.excel_process import convert_excel_to_images

load_dotenv()

dashscope.base_http_api_url = "https://dashscope.aliyuncs.com/api/v1"

# 读取Excel文件的前20行数据
def read_excel_first_20_rows(file_path):
    """
    读取Excel文件的前15行数据。

    参数：
        file_path (str): Excel文件的路径。

    返回：
        list: 包含前20行数据的列表，每行是一个包含单元格值的列表。
    """
    data = []

    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        # 获取第一个工作表
        sheet = workbook.active

        # 遍历前20行
        for row in sheet.iter_rows(min_row=1, max_row=15, values_only=True):
            data.append(list(row))

    except Exception as e:
        print(f"读取Excel文件时出错: {e}")

    return data

# 判断读取前20行数据的表头索引
def determine_header_index(rows):
    rows_str = json.dumps(rows, ensure_ascii=False)
    messages = [
        {"role": "system", "content": HEADER_ROW_DETECTION_PROMPT},
        {"role": "user", "content": rows_str},
    ]
    response = Generation.call(
        api_key=os.getenv("DASHSCOPE_API_KEY2"),
        model="qwen-flash",
        messages=messages,
        result_format="message",
        enable_thinking=False,
    )
    return response.get("output").choices[0].get("message").get("content")

# 按行切分Excel
def split_excel_by_rows_with_header(
        file_path, output_dir, header_index, rows_per_file=10
):
    """
    将Excel表格按行切分，每10行切分成一个新的Excel文件，且保留从第1行到表头索引行的所有数据。

    参数：
        file_path (str): 输入Excel文件的路径。
        output_dir (str): 输出文件的目录。
        header_index (int): 表头索引行的行号。
        rows_per_file (int): 每个文件包含的行数，默认为10。

    返回：
        None
    """
    import os
    import openpyxl

    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 获取总行数
        total_rows = sheet.max_row

        # 按行切分表格
        for start_row in range(header_index + 1, total_rows + 1, rows_per_file):
            # 创建新的工作簿
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = sheet.title

            # 复制从第1行到表头索引行的数据
            for row in sheet.iter_rows(
                    min_row=1, max_row=header_index, values_only=True
            ):
                new_sheet.append(row)

            # 复制当前分块的数据
            for row in sheet.iter_rows(
                    min_row=start_row,
                    max_row=min(start_row + rows_per_file - 1, total_rows),
                    values_only=True,
            ):
                new_sheet.append(row)

            # 保存新的工作簿
            output_file = os.path.join(
                output_dir,
                f"{sheet.title}_rows_{start_row}_to_{min(start_row + rows_per_file - 1, total_rows)}.xlsx",
            )
            new_workbook.save(output_file)
            print(f"已保存: {output_file}")

    except Exception as e:
        print(f"切分Excel文件时出错: {e}")

def format_excel_files_in_directory(directory):
    """
    给目录下的所有Excel文件添加全部框线，并调整列宽以防止数据被遮挡。

    参数：
        directory (str): 包含Excel文件的目录路径。

    返回：
        None
    """
    import os
    import openpyxl

    # 定义边框样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 遍历目录中的所有Excel文件
    file_paths = []
    for file_name in os.listdir(directory):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(directory, file_name)

            try:
                # 加载工作簿
                workbook = openpyxl.load_workbook(file_path)

                for sheet in workbook.worksheets:
                    # 调整列宽
                    for col in sheet.columns:
                        max_length = 0
                        col_letter = col[0].column_letter  # 获取列字母
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2  # 调整宽度
                        sheet.column_dimensions[col_letter].width = adjusted_width

                    # 添加边框
                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.border = thin_border

                # 保存修改
                workbook.save(file_path)
                file_paths.append(file_path)
                print(f"已格式化: {file_path}")
            except Exception as e:
                print(f"格式化文件时出错: {file_path}, 错误: {e}")
    return file_paths

if __name__ == "__main__":
    # 示例用法
    file_path = "../output_sheets/Sheet1 (2).xlsx"  # 替换为你的Excel文件路径
    rows = read_excel_first_20_rows(file_path)
    header_index = int(determine_header_index(rows))  # 获取表头索引
    print(f"表头索引: {header_index}")
    # output_dir = "output_excel"  # 替换为你的输出目录
    # split_excel_by_rows_with_header(file_path, output_dir, header_index + 1)
    directory = "output_excel"  # 替换为你的目录路径
    file_paths = format_excel_files_in_directory(directory)
    convert_excel_to_images(file_paths, "output_images")
