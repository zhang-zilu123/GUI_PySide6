import os
from docx2pdf import convert
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from openpyxl import load_workbook
import xlwings as xw

from spire.xls import *


def excel_to_pdf_1(excel_path, output_dir):
    # 注册支持中文的字体
    try:
        font_path = "C:/Windows/Fonts/simhei.ttf"
        pdfmetrics.registerFont(TTFont("SimHei", font_path))
    except:
        # 如果字体不存在，使用系统默认字体
        pass

    # 加载 Excel 文件
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 创建 PDF 文件
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_path = f"{output_dir}/output1.pdf"
    c = canvas.Canvas(output_path)

    try:
        c.setFont("SimHei", 12)
    except:
        # 如果SimHei字体不可用，使用默认字体
        c.setFont("Helvetica", 12)

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else "" for cell in row])
    text = c.beginText(40, 800)
    for row in data:
        line = " | ".join(row)
        text.textLine(line)
    c.drawText(text)
    c.save()

    return output_path


def excel_to_pdf_2(excel_path, output_path):
    workbook = Workbook()
    workbook.LoadFromFile(excel_path)
    workbook.SaveToFile(output_path, FileFormat.PDF)
    workbook.Dispose()
    return output_path


def excel_to_pdf_3(input_path, output_path=None):
    """
    使用 xlwings 将 Excel 转换为 PDF
    input_path: Excel 文件路径 (.xlsx / .xls)
    output_path: 输出 PDF 文件路径 (默认同目录同名)
    """
    if output_path is None:
        output_path = os.path.splitext(input_path)[0] + ".pdf"

    app = xw.App(visible=True)
    try:
        wb = app.books.open(input_path)
        # 导出为 PDF
        wb.api.ExportAsFixedFormat(0, output_path)  # 0 = xlTypePDF
        wb.close()
    finally:
        app.quit()

    return output_path


def word_to_pdf(input_path, output_path):
    convert(input_path, output_path)
    return output_path


if __name__ == '__main__':
    # 示例用法
    try:
        out = excel_to_pdf_3('./test.xlsx', './utils/test_out.pdf')
        print('转换成功，输出文件:', out)

    except Exception as e:
        print(f"转换失败: {e}")
